import datetime
from io import BytesIO
from typing import Dict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Side, Alignment, Border


def read_excel(file) -> Dict:
    """
    Parse uploaded file to dictionary
    :param file: uploaded file
    :return: dictionary with parsed data
    """
    def add_values(date_value: datetime.date):
        data[project_code][date_value] = (
            sheet.cell(row=row, column=col).value,
            sheet.cell(row=row, column=col + 1).value
        )

    # Мы, в том числе, оценивали умение преобразовывать данные с помощью pandas.
    # Грамотно используя его инструменты, можно было преобразовать данные лаконично в несколько строк кода.
    # Можно и без него обойтись, конечно.
    # Это не критично с точки зрения зачета тестового. Понятно, что требование для
    # бэкенд разраба весьма специфическое, и не стоит бросаться учить pandas.
    # Мы смотрим на это еще и как на умение решать новые задачи, изучать инструменты. В целом код понятный, ок.
    wb = load_workbook(filename=BytesIO(file.file.read()))
    sheet = wb['data']
    data = {}
    max_row = sheet.max_row
    max_column = sheet.max_column
    if sheet.max_row > 100002:
        max_row = 100002
    if sheet.max_column > 103:
        max_column = 103
    for row in range(3, max_row + 1):
        project_code = sheet.cell(row=row, column=1).value
        project_name = sheet.cell(row=row, column=2).value
        data[project_code] = {
            'name': project_name
        }
        for col in range(3, max_column, 2):
            date_value = sheet.cell(row=1, column=col).value
            if isinstance(date_value, datetime.datetime):
                add_values(date_value.date())
            elif isinstance(date_value, datetime.date):
                add_values(date_value)
            else:
                day, month, year = tuple(map(int, date_value.split('.')))
                date_value = datetime.date(year=year, month=month, day=day)
                add_values(date_value)
    return data


def write_excel(data: Dict) -> None:
    """
    Write data to Excel file
    :param data: parsed data
    :return: None
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'data'
    row = 3
    for project in data:
        sheet[f'A{row}'] = project
        sheet[f'B{row}'] = data[project]['name']
        col = 3
        for key in data[project]:
            sheet.cell(row=1, column=col, value=key)
            sheet.merge_cells(start_row=1, start_column=col,
                              end_row=1, end_column=col + 1)
            sheet.cell(row=2, column=col, value='план')
            sheet.cell(row=2, column=col + 1, value='факт')
            if key != 'name':
                sheet.cell(row=row, column=col, value=data[project][key][0])
                sheet.cell(row=row, column=col + 1, value=data[project][key][1])
                col += 2
        row += 1
    sheet['A2'] = 'Код'
    pf = PatternFill('solid', 'C0C0C0')
    thin = Side('thin', '000000')
    center = Alignment('center', 'center')
    for row in sheet.iter_rows(max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if cell.row < 3:
                cell.fill = pf
            cell.alignment = center
            cell.border = Border(top=thin, right=thin, left=thin, bottom=thin)
    sheet['B2'] = 'Наименование проекта'
    wb.save('test.xlsx')